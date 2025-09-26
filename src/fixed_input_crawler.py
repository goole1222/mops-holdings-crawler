#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复输入框定位的爬虫 - 针对正确的表单结构
"""

import time
import logging
import pandas as pd
import urllib3
import requests
import csv
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import os
import glob
from datetime import datetime
from typing import Optional

# 禁用SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class FixedInputCrawler:
    def __init__(self):
        """初始化修复输入框的爬虫"""
        self.setup_logging()
        self.driver = None
        self.all_data = {}
        self.failed_codes = []

        # 设置下载目录
        self.download_dir = os.path.join(os.getcwd(), "downloads")
        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir)
            self.logger.info(f"📁 创建下载目录: {self.download_dir}")

        # 导航信息
        self.main_url = "https://mops.twse.com.tw/mops/#/web/home"
        self.target_menu_text = "董監事持股餘額"

    def read_stock_codes(self, path="股票代號.txt"):
        """讀取股票代號清單文件"""
        import re
        codes = []
        try:
            with open(path, "r", encoding="utf-8") as f:
                for i, line in enumerate(f):
                    s = line.strip()
                    if not s or i == 0 and ("代號" in s or "code" in s.lower()):
                        continue
                    s = re.sub(r"[^\d]", "", s)  # 只留數字
                    if s:
                        codes.append(s)
        except Exception as e:
            self.logger.error(f"❌ 讀取代號清單失敗: {e}")
            return []

        # 去重，保留原順序
        seen, uniq = set(), []
        for c in codes:
            if c not in seen:
                seen.add(c)
                uniq.append(c)
        self.logger.info(f"📋 讀到 {len(uniq)} 個代號")
        return uniq

    def append_to_master_excel(self, out_path, df_chunk):
        """
        將 df_chunk（欄位必為 股票代號, 姓名, 目前持股）追加到 out_path 的「合併」工作表。
        若檔案不存在或沒有「合併」，就新建。
        """
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        import os

        df_chunk = df_chunk[["股票代號","姓名","目前持股"]].copy()

        if not os.path.exists(out_path):
            # 新建：直接寫入並套表頭樣式
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_chunk.to_excel(writer, sheet_name="合併", index=False)
                ws = writer.sheets["合併"]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 20
            self.logger.info(f"✅ 新建 Excel 並寫入: {out_path}")
            return

        # 已存在：讀舊合併、合併後整張重寫
        try:
            old = pd.read_excel(out_path, sheet_name="合併", engine="openpyxl")
        except Exception:
            old = pd.DataFrame(columns=["股票代號","姓名","目前持股"])

        merged = pd.concat([old, df_chunk], ignore_index=True)
        # 移除明顯表頭殘留與重複
        merged = merged.dropna(subset=["姓名"])
        merged = merged[~merged["姓名"].astype(str).str.contains("姓名|名稱", na=False)]
        merged = merged.drop_duplicates(subset=["股票代號","姓名"], keep="last")

        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            merged.to_excel(writer, sheet_name="合併", index=False)
            ws = writer.sheets["合併"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
        self.logger.info(f"✅ 追加資料至 Excel: {out_path} (目前共 {len(merged)} 筆)")

    def load_processed_codes(self, path="processed_codes.txt"):
        """載入已處理的代號清單"""
        import os
        if not os.path.exists(path):
            return set()
        with open(path, "r", encoding="utf-8") as f:
            return set([ln.strip() for ln in f if ln.strip()])

    def append_processed_code(self, code, path="processed_codes.txt"):
        """將已處理的代號追加到清單"""
        with open(path, "a", encoding="utf-8") as f:
            f.write(str(code).strip() + "\n")

    def setup_logging(self):
        """设置日志"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(message)s',
            handlers=[
                logging.FileHandler('fixed_input_crawler.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def setup_chrome(self):
        """设置Chrome选项"""
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument('--allow-running-insecure-content')
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--start-maximized')
        options.add_argument('--log-level=3')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])

        # 设置自动下载选项
        prefs = {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1,
            "profile.default_content_setting_values.popups": 1
        }
        options.add_experimental_option("prefs", prefs)

        return options

    def init_driver(self):
        """初始化浏览器驱动"""
        try:
            options = self.setup_chrome()
            self.driver = webdriver.Chrome(options=options)
            self.driver.set_page_load_timeout(30)
            self.driver.implicitly_wait(10)
            self.logger.info("✅ Chrome浏览器初始化成功")
            return True
        except Exception as e:
            self.logger.error(f"❌ 浏览器初始化失败: {e}")
            return False

    def navigate_to_target_page(self):
        """导航到董監事持股餘額页面"""
        try:
            # 步骤1: 进入主页
            self.logger.info(f"📖 步骤1: 访问主页 {self.main_url}")
            self.driver.get(self.main_url)
            time.sleep(5)

            self.logger.info(f"   页面标题: {self.driver.title}")

            # 步骤2: 点击董監事持股餘額菜单
            self.logger.info(f"🔍 步骤2: 寻找'{self.target_menu_text}'菜单")

            menu_selectors = [
                f"//a[contains(text(), '{self.target_menu_text}')]",
                f"//span[contains(text(), '{self.target_menu_text}')]",
                f"//div[contains(text(), '{self.target_menu_text}')]",
                f"//*[contains(text(), '{self.target_menu_text}')]"
            ]

            menu_element = None
            for selector in menu_selectors:
                try:
                    menu_elements = self.driver.find_elements(By.XPATH, selector)
                    for element in menu_elements:
                        if element.is_displayed() and element.is_enabled():
                            menu_element = element
                            self.logger.info(f"✅ 找到菜单项: {selector}")
                            break
                    if menu_element:
                        break
                except:
                    continue

            if not menu_element:
                self.logger.error(f"❌ 未找到'{self.target_menu_text}'菜单项")
                return False

            # 点击菜单
            self.logger.info(f"👆 点击'{self.target_menu_text}'菜单")
            try:
                menu_element.click()
            except:
                self.driver.execute_script("arguments[0].click();", menu_element)

            # 等待页面加载
            time.sleep(8)

            # 验证是否到达正确页面
            current_url = self.driver.current_url
            page_source = self.driver.page_source

            if "董監事持股餘額" in page_source or "查詢條件" in page_source:
                self.logger.info("✅ 成功导航到董監事持股餘額页面")
                return True
            else:
                self.logger.error("❌ 未能成功导航到目标页面")
                return False

        except Exception as e:
            self.logger.error(f"❌ 导航失败: {e}")
            return False

    def find_and_fill_company_input(self, stock_code):
        """寻找并填写公司代號或簡稱输入框"""
        try:
            self.logger.info(f"📝 步骤3: 寻找'公司代號或簡稱'输入框")

            # 等待页面完全加载
            time.sleep(3)

            # 多种策略寻找输入框
            input_strategies = [
                # 策略1: 通过标签文字寻找相邻的input
                "//label[contains(text(), '公司代號或簡稱')]/following-sibling::input",
                "//label[contains(text(), '公司代號或簡稱')]/..//input",

                # 策略2: 通过placeholder寻找
                "//input[contains(@placeholder, '1101')]",
                "//input[contains(@placeholder, '例如')]",

                # 策略3: 通过表单结构寻找
                "//div[contains(text(), '查詢條件')]//input",
                "//form//input[contains(@placeholder, '1101')]",

                # 策略4: 通过text内容寻找附近的input
                "//text()[contains(., '公司代號或簡稱')]/../..//input",
                "//*[contains(text(), '公司代號或簡稱')]/following::input[1]",
                "//*[contains(text(), '公司代號或簡稱')]/..//input",

                # 策略5: 查找表单中的所有text类型input
                "//div[contains(@class, 'form') or contains(@class, 'query')]//input[@type='text']",
                "//input[@type='text']",

                # 策略6: 通过name或id属性
                "//input[contains(@name, 'co_id') or contains(@name, 'company')]",
                "//input[contains(@id, 'co_id') or contains(@id, 'company')]"
            ]

            input_element = None
            found_strategy = None

            for i, strategy in enumerate(input_strategies, 1):
                try:
                    self.logger.info(f"   尝试策略{i}: {strategy[:50]}...")
                    elements = self.driver.find_elements(By.XPATH, strategy)

                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            # 验证这是否是正确的输入框
                            placeholder = element.get_attribute('placeholder') or ""
                            name = element.get_attribute('name') or ""
                            id_attr = element.get_attribute('id') or ""

                            self.logger.info(f"   找到输入框: placeholder='{placeholder}', name='{name}', id='{id_attr}'")

                            # 如果placeholder包含'1101'或其他相关关键词，这很可能是正确的输入框
                            if ('1101' in placeholder or '例如' in placeholder or
                                'co_id' in name.lower() or 'company' in name.lower()):
                                input_element = element
                                found_strategy = strategy
                                self.logger.info(f"✅ 找到目标输入框！使用策略{i}")
                                break

                            # 如果没有明确标识，但是在查询表单中，也可能是正确的
                            if not input_element:
                                input_element = element
                                found_strategy = strategy
                                self.logger.info(f"✅ 找到可能的输入框，使用策略{i}")

                    if input_element:
                        break

                except Exception as e:
                    self.logger.info(f"   策略{i}失败: {e}")
                    continue

            if not input_element:
                self.logger.error("❌ 所有策略都未找到输入框")

                # 调试信息：显示页面中所有input元素
                self.logger.info("🔍 调试信息 - 页面中所有input元素:")
                try:
                    all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
                    for j, inp in enumerate(all_inputs):
                        try:
                            if inp.is_displayed():
                                placeholder = inp.get_attribute('placeholder') or ""
                                name = inp.get_attribute('name') or ""
                                type_attr = inp.get_attribute('type') or ""
                                self.logger.info(f"   Input{j+1}: type='{type_attr}', placeholder='{placeholder}', name='{name}'")
                        except:
                            pass
                except:
                    pass

                return False

            # 填写股票代号
            self.logger.info(f"📝 在输入框中输入股票代号: {stock_code}")

            try:
                # 清空输入框
                input_element.clear()
                time.sleep(1)

                # 输入股票代号
                input_element.send_keys(stock_code)
                time.sleep(1)

                # 验证输入
                current_value = input_element.get_attribute('value')
                if current_value == stock_code:
                    self.logger.info(f"✅ 股票代号输入成功: {current_value}")
                    return True
                else:
                    self.logger.warning(f"⚠️ 输入验证失败: 期望'{stock_code}', 实际'{current_value}'")
                    return True  # 仍然继续，可能是显示延迟

            except Exception as e:
                self.logger.error(f"❌ 输入股票代号失败: {e}")
                return False

        except Exception as e:
            self.logger.error(f"❌ 寻找输入框过程失败: {e}")
            return False

    def click_query_button(self):
        """点击查询按钮"""
        try:
            self.logger.info("🔍 步骤4: 寻找并点击'查詢'按钮")

            # 多种策略寻找查询按钮
            button_strategies = [
                # 策略1: 直接查找查询按钮
                "//button[contains(text(), '查詢')]",
                "//input[@value='查詢']",
                "//button[contains(@class, 'query') or contains(@class, 'search')]",

                # 策略2: 在表单中查找提交按钮
                "//form//button[@type='submit']",
                "//form//input[@type='submit']",

                # 策略3: 查找蓝色或主要按钮
                "//button[contains(@class, 'btn-primary') or contains(@class, 'primary')]",
                "//button[contains(@class, 'blue') or contains(@style, 'blue')]",

                # 策略4: 在查询条件附近查找按钮
                "//div[contains(text(), '查詢條件')]//button",
                "//*[contains(text(), '查詢')]"
            ]

            button_element = None
            for i, strategy in enumerate(button_strategies, 1):
                try:
                    self.logger.info(f"   尝试按钮策略{i}: {strategy[:50]}...")
                    elements = self.driver.find_elements(By.XPATH, strategy)

                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            button_text = element.text or element.get_attribute('value') or ""
                            self.logger.info(f"   找到按钮: 文字='{button_text}'")

                            if '查詢' in button_text or element.get_attribute('type') == 'submit':
                                button_element = element
                                self.logger.info(f"✅ 找到查询按钮！使用策略{i}")
                                break

                    if button_element:
                        break

                except Exception as e:
                    self.logger.info(f"   按钮策略{i}失败: {e}")
                    continue

            if not button_element:
                self.logger.error("❌ 未找到查询按钮")
                return False

            # 点击查询按钮
            self.logger.info("👆 点击查询按钮")
            try:
                button_element.click()
                self.logger.info("   直接点击成功")
            except:
                try:
                    self.driver.execute_script("arguments[0].click();", button_element)
                    self.logger.info("   JavaScript点击成功")
                except:
                    self.logger.error("   所有点击方式都失败")
                    return False

            # 等待查询结果
            self.logger.info("⏳ 等待查询结果加载...")
            time.sleep(8)

            return True

        except Exception as e:
            self.logger.error(f"❌ 点击查询按钮失败: {e}")
            return False

    def _requests_session_from_driver(self):
        """將 Selenium cookies 轉成 requests 可用的 session"""
        s = requests.Session()
        for c in self.driver.get_cookies():
            s.cookies.set(c["name"], c["value"], domain=c.get("domain"))
        # 帶上 UA
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        return s

    def clear_old_downloads(self):
        """清除下载目录中的旧CSV文件"""
        try:
            csv_files = glob.glob(os.path.join(self.download_dir, "*.csv"))
            crdownload_files = glob.glob(os.path.join(self.download_dir, "*.crdownload"))

            all_files_to_remove = csv_files + crdownload_files
            removed_count = 0

            for file_path in all_files_to_remove:
                try:
                    os.remove(file_path)
                    removed_count += 1
                    self.logger.info(f"🗑️ 已删除旧文件: {os.path.basename(file_path)}")
                except Exception as e:
                    self.logger.warning(f"⚠️ 删除文件失败 {os.path.basename(file_path)}: {e}")

            if removed_count > 0:
                self.logger.info(f"🧹 清理完成，共删除 {removed_count} 个文件")
            else:
                self.logger.info("ℹ️ 下载目录中无需清理的文件")

        except Exception as e:
            self.logger.error(f"❌ 清理下载目录失败: {e}")

    def download_csv_and_parse(self):
        import time, os, glob
        from datetime import datetime
        self.logger.info("📥 步骤4a: 嘗試下載CSV檔案")
        self.clear_old_downloads()
        time.sleep(1)

        # 1) 先找 a[href*=.csv] 或 下載CSV 按鈕
        candidates = []
        xpaths = [
            "//a[contains(@href,'.csv')]",
            "//button[contains(normalize-space(.),'下載CSV')]",
            "//span[contains(normalize-space(.),'下載CSV')]/ancestor::a",
            "//span[contains(normalize-space(.),'下載CSV')]/ancestor::button"
        ]
        for xp in xpaths:
            els = self.driver.find_elements(By.XPATH, xp)
            for el in els:
                if el.is_displayed() and el.is_enabled():
                    candidates.append(el)
        if not candidates:
            self.logger.warning("⚠️ 未找到CSV下載元素")
            return None

        # 2) 有 href 的話，直接用 requests 下載
        for el in candidates:
            href = el.get_attribute("href")
            if href and ".csv" in href.lower():
                try:
                    self.logger.info(f"🔗 直接請求 CSV: {href[:120]}...")
                    sess = self._requests_session_from_driver()
                    r = sess.get(href, timeout=20)
                    content_type = r.headers.get("Content-Type", "").lower()
                    self.logger.info(f"📄 回應 Content-Type: {content_type}, 內容長度: {len(r.content)} bytes")

                    # 如果回應成功且有內容，就嘗試解析 (不限制 Content-Type)
                    if r.status_code == 200 and len(r.content) > 0:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        csv_path = os.path.join(self.download_dir, f"mops_{ts}.csv")
                        with open(csv_path, "wb") as f:
                            f.write(r.content)
                        self.logger.info(f"✅ 以 requests 下載檔案成功: {os.path.basename(csv_path)}")
                        return self._read_and_filter_csv(csv_path)
                except Exception as e:
                    self.logger.warning(f"⚠️ 直接請求 CSV 失敗: {e}")

        # 3) 沒有 href（或 requests 失敗）→ 退回點擊 + 目錄監看
        try:
            target = candidates[0]
            self.driver.execute_script("arguments[0].scrollIntoView(true);", target)
            time.sleep(0.5)
            try:
                self.driver.execute_script("arguments[0].click();", target)
            except:
                target.click()
            self.logger.info("🖱️ 已點擊下載CSV按鈕，開始監看下載資料夾（30秒）...")
        except Exception as e:
            self.logger.error(f"❌ 點擊下載CSV失敗: {e}")
            return None

        # 4) 監看 30 秒，接受任意新檔（非 .crdownload，包含無副檔名檔案）
        latest_file = None
        start = time.time()
        while time.time() - start < 30:
            time.sleep(1)
            files = [f for f in glob.glob(os.path.join(self.download_dir, "*")) if not f.endswith(".crdownload")]
            if files:
                latest = max(files, key=os.path.getmtime)
                # 避免把很舊的檔當成新下載
                if time.time() - os.path.getmtime(latest) < 60:
                    latest_file = latest
                    self.logger.info(f"📁 偵測到新檔案: {os.path.basename(latest_file)} (檔案大小: {os.path.getsize(latest_file)} bytes)")
                    break
        if not latest_file:
            self.logger.warning("⚠️ 下載超時或無新檔，放棄 CSV 流程")
            return None

        self.logger.info(f"✅ 使用 Python engine + on_bad_lines='skip' 解析檔案")
        return self._read_and_filter_csv(latest_file)

    def _read_and_filter_csv(self, path):
        import csv
        import re
        import pandas as pd

        # ---- A. 讀檔 & 編碼容錯 ----
        encodings = ["utf-8-sig", "utf-8", "big5", "cp950"]
        text = None
        last_err = None
        for enc in encodings:
            try:
                with open(path, "r", encoding=enc, errors="replace") as f:
                    text = f.read()
                if text:
                    self.logger.info(f"🔤 成功以 {enc} 編碼讀取檔案")
                    break
            except Exception as e:
                last_err = e
                continue
        if not text:
            self.logger.error(f"❌ 無法以常見編碼讀取檔案: {last_err}")
            return None

        # 清理 BOM/奇怪控制字元
        text = text.replace("\ufeff", "")
        text = text.replace("\x00", "")

        # ---- B. 尋找真正表頭所在行（包含「姓名」與「目前持股」關鍵詞的行） ----
        lines = [ln.strip() for ln in text.splitlines() if ln.strip() != ""]
        header_idx = None
        header_line = None
        for i, ln in enumerate(lines[:100]):  # 只在前100行找表頭
            if ("姓名" in ln or "名稱" in ln) and ("目前持股" in ln  or "目前持股(股)" in ln):
                header_idx = i
                header_line = ln
                self.logger.info(f"🎯 在第{i+1}行找到表頭: {ln[:100]}")
                break
        if header_idx is None:
            # 若找不到，退而求其次：找到第一行至少包含兩個欄位分隔符的行
            for i, ln in enumerate(lines[:100]):
                if ln.count(",") >= 1 or ln.count(";") >= 1 or ln.count("\t") >= 1:
                    header_idx = i
                    header_line = ln
                    self.logger.info(f"📊 在第{i+1}行找到可能的表頭: {ln[:100]}")
                    break
        if header_idx is None:
            self.logger.error("❌ 找不到表頭列，無法解析 CSV")
            return None

        data_str = "\n".join(lines[header_idx:])

        # ---- C. 自動偵測分隔符號 ----
        dialect = None
        try:
            dialect = csv.Sniffer().sniff(data_str.splitlines()[0] + "\n" + data_str.splitlines()[1])
            delimiter = dialect.delimiter
            self.logger.info(f"🔍 自動偵測分隔符: '{delimiter}'")
        except Exception:
            # 手動猜測：優先 , ; \t
            if header_line.count(",") >= 1:
                delimiter = ","
            elif header_line.count(";") >= 1:
                delimiter = ";"
            elif header_line.count("\t") >= 1:
                delimiter = "\t"
            else:
                # 最後保底：逗號
                delimiter = ","
            self.logger.info(f"🔍 手動判斷分隔符: '{delimiter}'")

        # ---- D. 多策略讀取 pandas ----
        candidates = []
        # 策略1：讓 pandas 自動判斷（python engine + on_bad_lines='skip'）
        try:
            df1 = pd.read_csv(
                pd.io.common.StringIO(data_str),
                sep=delimiter if delimiter != "," else None,  # 逗號時讓 sniff 決定
                engine="python",
                on_bad_lines="skip"
            )
            if df1 is not None and not df1.empty:
                candidates.append(df1)
                self.logger.info(f"✅ 策略1成功: {len(df1)} 行 × {len(df1.columns)} 欄")
        except Exception as e:
            self.logger.warning(f"⚠️ 讀取策略1失敗: {e}")

        # 策略2：明確指定 sep
        try:
            df2 = pd.read_csv(pd.io.common.StringIO(data_str), sep=delimiter, engine="python", on_bad_lines="skip")
            if df2 is not None and not df2.empty:
                candidates.append(df2)
                self.logger.info(f"✅ 策略2成功: {len(df2)} 行 × {len(df2.columns)} 欄")
        except Exception as e:
            self.logger.warning(f"⚠️ 讀取策略2失敗: {e}")

        # 策略3：若欄位仍亂，先用 csv.reader 解析成 rows，再手工變 DataFrame
        if not candidates:
            try:
                reader = csv.reader(pd.io.common.StringIO(data_str), delimiter=delimiter)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
                # 以第一列為欄名（若第一列不是中文欄名，後面會再重命名）
                if len(rows) >= 2:
                    header = rows[0]
                    body = rows[1:]
                    df3 = pd.DataFrame(body, columns=header)
                    candidates.append(df3)
                    self.logger.info(f"✅ 策略3成功: {len(df3)} 行 × {len(df3.columns)} 欄")
            except Exception as e:
                self.logger.warning(f"⚠️ 讀取策略3失敗: {e}")

        if not candidates:
            self.logger.error("❌ 所有 CSV 讀取策略均失敗")
            return None

        # ---- E. 嘗試從候選 df 中挑出含關鍵欄位者 ----
        def _pick_columns(df):
            cols = list(map(str, df.columns))
            # 清理欄名空白
            df = df.rename(columns={c: str(c).strip() for c in df.columns})

            name_col = next((c for c in df.columns if any(k in str(c) for k in ["姓名","名稱","姓名/名稱","董監事姓名"])), None)
            hold_col = next((c for c in df.columns if any(k in str(c) for k in ["目前持股","目前持股數","目前持股(股)","現有持股"])), None)

            # 若第一列其實是表頭，欄名在第一列內容，再提升一行為欄名
            if not name_col or not hold_col:
                if len(df) >= 1:
                    first_row = df.iloc[0].astype(str).tolist()
                    if any("姓名" in x or "名稱" in x for x in first_row):
                        df2 = df[1:].copy()
                        df2.columns = first_row
                        df = df2
                        name_col = next((c for c in df.columns if any(k in str(c) for k in ["姓名","名稱","姓名/名稱","董監事姓名"])), None)
                        hold_col = next((c for c in df.columns if any(k in str(c) for k in ["目前持股","目前持股數","目前持股(股)","現有持股"])), None)

            return df, name_col, hold_col

        chosen = None
        for i, df in enumerate(candidates):
            df2, name_col, hold_col = _pick_columns(df)
            self.logger.info(f"📋 候選{i+1}: 姓名欄='{name_col}', 持股欄='{hold_col}'")
            if name_col and hold_col:
                chosen = (df2, name_col, hold_col)
                self.logger.info(f"✅ 選中候選{i+1}")
                break

        if not chosen:
            # 再嘗試把所有欄名/內容去空白後重試一次
            for i, df in enumerate(candidates):
                df.columns = [str(c).strip() for c in df.columns]
                df2, name_col, hold_col = _pick_columns(df)
                if name_col and hold_col:
                    chosen = (df2, name_col, hold_col)
                    self.logger.info(f"✅ 清理空白後選中候選{i+1}")
                    break

        if not chosen:
            self.logger.error("❌ 無法識別「姓名」與「目前持股」欄位")
            return None

        df, name_col, hold_col = chosen
        out = df[[name_col, hold_col]].copy()
        out.columns = ["姓名", "目前持股"]

        # 去掉明顯的表頭/空白列
        out = out.dropna(subset=["姓名"])
        out = out[~out["姓名"].astype(str).str.contains("姓名|名稱")]
        out["目前持股"] = out["目前持股"].astype(str).str.strip()
        out = out.drop_duplicates(subset=["姓名"])

        self.logger.info(f"✅ CSV 數據處理完成：{len(out)} 筆")
        return out if not out.empty else None

    def extract_data_from_divs(self, stock_code):
        """從 div/span 區塊提取姓名和目前持股數據"""
        try:
            self.logger.info(f"🔍 嘗試從 div/span 區塊提取股票 {stock_code} 的數據")

            # 等待數據完全加載
            time.sleep(3)

            # 尋找所有包含「姓名：」的元素
            name_elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), '姓名：')]")
            self.logger.info(f"📋 找到 {len(name_elements)} 個包含「姓名：」的元素")

            if not name_elements:
                self.logger.info("ℹ️ 未找到包含「姓名：」的元素，將嘗試表格解析")
                return None

            extracted_data = []

            for name_element in name_elements:
                try:
                    # 提取姓名
                    name_text = name_element.text.strip()
                    if "姓名：" in name_text:
                        name = name_text.split("姓名：")[1].strip()
                        if not name:
                            continue

                        # 尋找對應的「目前持股：」元素
                        holdings = None

                        # 策略1: 在同一個父元素中尋找
                        parent = name_element.find_element(By.XPATH, "./..")
                        holdings_elements = parent.find_elements(By.XPATH, ".//*[contains(text(), '目前持股：')]")

                        if holdings_elements:
                            holdings_text = holdings_elements[0].text.strip()
                            if "目前持股：" in holdings_text:
                                holdings = holdings_text.split("目前持股：")[1].strip()
                        else:
                            # 策略2: 在整個頁面中尋找緊鄰的「目前持股：」元素
                            following_elements = name_element.find_elements(By.XPATH, "./following::*[contains(text(), '目前持股：')][1]")
                            if following_elements:
                                holdings_text = following_elements[0].text.strip()
                                if "目前持股：" in holdings_text:
                                    holdings = holdings_text.split("目前持股：")[1].strip()
                            else:
                                # 策略3: 尋找下一個兄弟元素或鄰近元素
                                siblings = name_element.find_elements(By.XPATH, "./following-sibling::*")
                                for sibling in siblings[:5]:  # 只檢查前5個兄弟元素
                                    if "目前持股：" in sibling.text:
                                        holdings_text = sibling.text.strip()
                                        holdings = holdings_text.split("目前持股：")[1].strip()
                                        break

                        if name and holdings is not None:
                            extracted_data.append({
                                "姓名": name,
                                "目前持股": holdings
                            })
                            self.logger.info(f"   ✅ 提取成功: 姓名={name}, 目前持股={holdings}")
                        else:
                            self.logger.info(f"   ⚠️ 無法找到對應的持股數據: 姓名={name}")

                except Exception as e:
                    self.logger.info(f"   ⚠️ 處理元素時出錯: {e}")
                    continue

            if extracted_data:
                df = pd.DataFrame(extracted_data)
                self.logger.info(f"✅ 從 div/span 成功提取 {len(extracted_data)} 行數據")
                return df
            else:
                self.logger.info("ℹ️ div/span 區塊中未提取到有效數據")
                return None

        except Exception as e:
            self.logger.error(f"❌ div/span 數據提取失敗: {e}")
            return None

    def extract_data_from_table(self, stock_code):
        """從表格提取姓名和目前持股數據（原始邏輯）"""
        try:
            self.logger.info(f"📊 從表格提取股票 {stock_code} 的數據")

            # 寻找数据表格
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            self.logger.info(f"📋 页面中找到 {len(tables)} 个表格")

            if not tables:
                self.logger.warning("⚠️ 页面中没有找到表格")
                return None

            # 选择最佳表格
            target_table = None
            max_score = 0

            for i, table in enumerate(tables):
                try:
                    table_text = table.text
                    keywords = ["姓名", "持股", "董事", "監事", "目前", "現任"]
                    score = sum(1 for keyword in keywords if keyword in table_text)

                    rows = table.find_elements(By.TAG_NAME, "tr")
                    if len(rows) < 2:
                        score = 0

                    self.logger.info(f"   表格{i+1}: 评分{score}, 行数{len(rows)}")

                    if score > max_score and score >= 2:
                        max_score = score
                        target_table = table

                except:
                    continue

            if not target_table:
                self.logger.warning("⚠️ 未找到包含股东数据的表格")
                return None

            # 提取表格数据
            rows = target_table.find_elements(By.TAG_NAME, "tr")
            headers = []
            data_rows = []

            for row in rows:
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        cells = row.find_elements(By.TAG_NAME, "th")
                        if cells and not headers:
                            headers = [cell.text.strip() for cell in cells if cell.text.strip()]

                    if len(cells) >= 2 and any(cell.text.strip() for cell in cells):
                        row_data = [cell.text.strip() for cell in cells]
                        data_rows.append(row_data)

                except:
                    continue

            if not data_rows:
                self.logger.warning("⚠️ 没有提取到有效数据")
                return None

            self.logger.info(f"📊 表头: {headers}")
            self.logger.info(f"📊 数据行: {len(data_rows)}")

            # 智能识别姓名和持股列
            name_col_index = None
            holdings_col_index = None

            # 寻找姓名列
            for i, header in enumerate(headers):
                if any(keyword in header for keyword in ["姓名", "名稱"]):
                    name_col_index = i
                    break
            if name_col_index is None and len(headers) >= 2:
                name_col_index = 1

            # 寻找持股列
            for i, header in enumerate(headers):
                if any(keyword in header for keyword in ["目前持股", "目前"]):
                    holdings_col_index = i
                    break
            if holdings_col_index is None:
                for col in range(2, min(len(headers), 6)):
                    has_numbers = False
                    for row in data_rows[:3]:
                        if col < len(row):
                            cell_text = row[col].replace(',', '').replace(' ', '')
                            if cell_text.isdigit() and len(cell_text) > 2:
                                has_numbers = True
                                break
                    if has_numbers:
                        holdings_col_index = col
                        break

            if name_col_index is None or holdings_col_index is None:
                self.logger.error(f"❌ 无法识别姓名列({name_col_index})或持股列({holdings_col_index})")
                return None

            # 提取数据
            extracted_data = []
            for row in data_rows:
                try:
                    if name_col_index < len(row) and holdings_col_index < len(row):
                        name = row[name_col_index].strip()
                        holdings = row[holdings_col_index].strip()

                        if (name and holdings and name not in ["姓名", "名稱"] and
                            not any(keyword in name for keyword in ["職稱", "姓名"])):
                            extracted_data.append({
                                "姓名": name,
                                "目前持股": holdings
                            })
                except:
                    continue

            if extracted_data:
                df = pd.DataFrame(extracted_data)
                self.logger.info(f"✅ 從表格成功提取 {len(extracted_data)} 行数据")
                return df
            else:
                self.logger.warning("⚠️ 表格中没有提取到有效的姓名和持股数据")
                return None

        except Exception as e:
            self.logger.error(f"❌ 表格数据提取失败: {e}")
            return None

    def extract_name_and_holdings_data(self, stock_code):
        """提取姓名和目前持股数据（優先使用 div/span，其次使用表格）"""
        try:
            self.logger.info(f"📊 步骤5: 提取股票 {stock_code} 的姓名和持股数据")

            # 等待数据完全加载
            time.sleep(3)

            # 验证是否有查询结果
            page_source = self.driver.page_source
            if stock_code not in page_source and "股份有限公司" not in page_source:
                self.logger.warning("⚠️ 页面中可能没有查询结果")

            # 優先嘗試從 div/span 區塊提取數據
            div_data = self.extract_data_from_divs(stock_code)
            if div_data is not None and len(div_data) > 0:
                self.logger.info("✅ 成功從 div/span 區塊提取數據")
                return div_data

            # 如果 div/span 提取失敗，回到原始的表格提取邏輯
            self.logger.info("📋 div/span 提取無數據，嘗試表格提取")
            table_data = self.extract_data_from_table(stock_code)
            if table_data is not None and len(table_data) > 0:
                self.logger.info("✅ 成功從表格提取數據")
                return table_data

            self.logger.warning("⚠️ 所有提取方式都未能獲得有效數據")
            return None

        except Exception as e:
            self.logger.error(f"❌ 数据提取失败: {e}")
            return None

    def process_single_stock(self, stock_code):
        """处理单个股票的完整流程"""
        try:
            self.logger.info(f"\n{'='*60}")
            self.logger.info(f"📈 开始处理股票: {stock_code}")
            self.logger.info(f"{'='*60}")

            # 导航到目标页面
            if not self.navigate_to_target_page():
                return False

            # 寻找并填写输入框
            if not self.find_and_fill_company_input(stock_code):
                return False

            # 点击查询按钮
            if not self.click_query_button():
                return False

            # 數據提取（優先順序：CSV下載 > div/span解析 > 表格解析）
            data = None

            # 步骤4a: 优先尝试CSV下载
            data = self.download_csv_and_parse()
            if data is not None and len(data) > 0:
                # 在存入 self.all_data 之前，加入股票代號欄位（如果尚未插入）
                if "股票代號" not in data.columns:
                    data.insert(0, "股票代號", stock_code)
                self.all_data[stock_code] = data
                self.logger.info(f"✅ 股票 {stock_code} 通過CSV下載處理成功")
                return True

            # 步骤4b: 如果CSV下载失败，退回到原有的解析逻辑
            self.logger.info("📋 CSV下載失敗，使用備援解析方式")
            data = self.extract_name_and_holdings_data(stock_code)
            if data is not None and len(data) > 0:
                # 在存入 self.all_data 之前，加入股票代號欄位（如果尚未插入）
                if "股票代號" not in data.columns:
                    data.insert(0, "股票代號", stock_code)
                self.all_data[stock_code] = data
                self.logger.info(f"✅ 股票 {stock_code} 通過備援解析處理成功")
                return True
            else:
                self.logger.error(f"❌ 股票 {stock_code} 所有數據提取方式都失敗")
                return False

        except Exception as e:
            self.logger.error(f"❌ 处理股票 {stock_code} 异常: {e}")
            return False

    def save_to_excel(self, output_path, make_per_sheet=False):
        """保存到Excel"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 合併表
                merged = []
                for stock_code, df in self.all_data.items():
                    merged.append(df[["股票代號", "姓名", "目前持股"]].copy())

                if merged:
                    merged_df = pd.concat(merged, ignore_index=True)
                    merged_df.to_excel(writer, sheet_name="合併", index=False)

                    # 簡單格式
                    ws = writer.sheets["合併"]
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions['A'].width = 14  # 股票代號
                    ws.column_dimensions['B'].width = 30  # 姓名
                    ws.column_dimensions['C'].width = 20  # 目前持股

                # 依代號各自一張分頁（可選）
                if make_per_sheet:
                    for stock_code, data in self.all_data.items():
                        sheet_name = f"股票_{stock_code}"[:31]  # Excel 名稱長度限制
                        data.to_excel(writer, sheet_name=sheet_name, index=False)

                        # 格式化
                        worksheet = writer.sheets[sheet_name]
                        from openpyxl.styles import Font, PatternFill, Alignment

                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                        for cell in worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center")

                        worksheet.column_dimensions['A'].width = 14  # 股票代號
                        worksheet.column_dimensions['B'].width = 30  # 姓名
                        worksheet.column_dimensions['C'].width = 20  # 目前持股

                if self.failed_codes:
                    pd.DataFrame({"失敗的股票代號": self.failed_codes}).to_excel(writer, sheet_name="失敗記錄", index=False)

            self.logger.info(f"✅ Excel文件保存成功: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"❌ Excel保存失败: {e}")
            return False

    def run_batch(self, codes_file="股票代號.txt", make_per_sheet=False, throttle_sec=1.5, retry=1):
        """批次處理股票清單"""
        try:
            self.logger.info("="*80)
            self.logger.info("🚀 批次抓取開始")
            self.logger.info("="*80)

            if not self.init_driver():
                return False

            codes = self.read_stock_codes(codes_file)
            if not codes:
                self.logger.error("❌ 沒有可用的代號")
                return False

            for idx, code in enumerate(codes, 1):
                ok = False
                for r in range(retry + 1):
                    self.logger.info(f"\n[{idx}/{len(codes)}] ▶︎ 處理 {code}（重試 {r}/{retry}）")
                    ok = self.process_single_stock(code)
                    if ok:
                        break
                    time.sleep(2)
                if not ok:
                    self.failed_codes.append(code)
                time.sleep(throttle_sec)  # 節流，避免過快

            if self.all_data:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out = f"董監事持股_批次_{ts}.xlsx"
                self.save_to_excel(out, make_per_sheet=make_per_sheet)
                self.logger.info(f"🎯 批次完成！成功 {len(self.all_data)} 檔，失敗 {len(self.failed_codes)} 檔")
                return True
            else:
                self.logger.error("❌ 無任何成功資料")
                return False

        except Exception as e:
            self.logger.error(f"❌ 批次執行發生例外: {e}")
            return False
        finally:
            if self.driver:
                self.driver.quit()

    def run_batch_resume(self, codes_file="股票代號.txt", out_path=None, throttle_sec=1.5, retry=1):
        """批次處理股票清單（可續跑版本）"""
        import os
        from datetime import datetime

        self.logger.info("="*80)
        self.logger.info("🚀 批次抓取（可續跑）開始")
        self.logger.info("="*80)

        if not self.init_driver():
            return False

        codes = self.read_stock_codes(codes_file)
        if not codes:
            self.logger.error("❌ 沒有可用的代號")
            return False

        done = self.load_processed_codes()
        pending = [c for c in codes if c not in done]
        self.logger.info(f"✅ 已完成 {len(done)} 檔，待處理 {len(pending)} 檔")

        if out_path is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = f"董監事持股_合併_{ts}.xlsx"

        success_cnt = 0
        for idx, code in enumerate(pending, 1):
            ok = False
            for r in range(retry + 1):
                self.logger.info(f"[{idx}/{len(pending)}] ▶︎ {code}（重試 {r}/{retry}）")
                ok = self.process_single_stock(code)  # 內含 CSV/備援解析
                if ok and code in self.all_data:
                    # 立刻寫入 Excel（合併表），並標記 processed
                    df = self.all_data[code][["股票代號","姓名","目前持股"]].copy()
                    self.append_to_master_excel(out_path, df)
                    self.append_processed_code(code)
                    # 釋放該代號的暫存以省記憶體
                    del self.all_data[code]
                    success_cnt += 1
                    break
                time.sleep(2)

            if not ok:
                self.failed_codes.append(code)

            time.sleep(throttle_sec)

        # 最後把失敗清單寫入到同一份 Excel 的「失敗記錄」sheet（覆蓋或新建）
        try:
            import pandas as pd
            mode = "a" if os.path.exists(out_path) else "w"
            with pd.ExcelWriter(out_path, engine="openpyxl", mode="a" if mode=="a" else "w") as writer:
                if self.failed_codes:
                    pd.DataFrame({"失敗的股票代號": self.failed_codes}).to_excel(writer, sheet_name="失敗記錄", index=False)
        except Exception as e:
            self.logger.warning(f"⚠️ 寫入失敗記錄時發生例外：{e}")

        self.logger.info(f"🎯 完成：成功 {success_cnt} 檔，失敗 {len(self.failed_codes)} 檔；輸出：{out_path}")
        if self.driver:
            self.driver.quit()
        return success_cnt > 0

    def run_fixed_test(self, stock_codes=['1235']):
        """运行修复版测试"""
        try:
            self.logger.info("=" * 80)
            self.logger.info("🔧 修复输入框定位的爬虫测试")
            self.logger.info("=" * 80)

            if not self.init_driver():
                return False

            for stock_code in stock_codes:
                success = self.process_single_stock(stock_code)
                if not success:
                    self.failed_codes.append(stock_code)
                time.sleep(3)

            if self.all_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"修复版股票数据_{timestamp}.xlsx"
                self.save_to_excel(output_file)

                self.logger.info(f"\n🎯 处理完成!")
                self.logger.info(f"   成功: {len(self.all_data)} 个股票")
                self.logger.info(f"   失败: {len(self.failed_codes)} 个股票")

            return len(self.all_data) > 0

        except Exception as e:
            self.logger.error(f"❌ 测试运行异常: {e}")
            return False
        finally:
            if self.driver:
                self.driver.quit()

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--codes-file", default="股票代號.txt")
    parser.add_argument("--out", default=None, help="輸出 Excel 路徑；不填則自動依時間命名")
    parser.add_argument("--retry", type=int, default=1)
    parser.add_argument("--throttle", type=float, default=1.5)
    args = parser.parse_args()

    print("🔧 股票爬蟲（邊跑邊寫・可續跑）")
    print("="*50)
    crawler = FixedInputCrawler()
    ok = crawler.run_batch_resume(
        codes_file=args.codes_file,
        out_path=args.out,
        throttle_sec=args.throttle,
        retry=args.retry
    )
    print("\n✅ 完成" if ok else "\n❌ 失敗，請看 log")

if __name__ == "__main__":
    main()